"""Exportación reutilizable a Excel/PDF para vistas basadas en funciones (FBV).

Mismo estilo visual que ExportListMixin (usado por los CRUD genéricos), pero
invocable desde una FBV pasando encabezados + filas ya resueltas.

Uso típico en una vista:
    export = request.GET.get('export')
    if export == 'excel':
        return export_excel(headers, rows, 'facturas', 'Listado de Facturas')
    if export == 'pdf':
        return export_pdf(headers, rows, 'facturas', 'Listado de Facturas')
"""
from django.http import HttpResponse
from django.utils import timezone


def _fname(filename, ext):
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
    return f'{filename}_{stamp}.{ext}'


def export_excel(headers, rows, filename='export', title=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = (title or filename)[:31]

    ws.append(list(headers))
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='343A40')
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill

    widths = [len(str(h)) for h in headers]
    for row in rows:
        clean = ['' if v is None else v for v in row]
        ws.append(clean)
        for i, v in enumerate(clean):
            widths[i] = max(widths[i], len(str(v)))

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(60, max(10, w + 2))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{_fname(filename, "xlsx")}"'
    wb.save(response)
    return response


def export_pdf(headers, rows, filename='export', title=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape, portrait
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_fname(filename, "pdf")}"'

    headers = list(headers)
    ncols = len(headers) or 1

    pagesize = portrait(A4) if ncols <= 4 else landscape(A4)
    if ncols <= 5:
        font_size = 9
    elif ncols <= 8:
        font_size = 7
    else:
        font_size = 6

    margin = 1 * cm
    doc = SimpleDocTemplate(
        response, pagesize=pagesize,
        leftMargin=margin, rightMargin=margin, topMargin=margin, bottomMargin=margin,
    )

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('cell', parent=styles['Normal'], fontSize=font_size, leading=font_size + 2)
    head_style = ParagraphStyle('head', parent=cell_style, textColor=colors.white,
                                fontName='Helvetica-Bold')
    elements = [
        Paragraph(title or filename, styles['Title']),
        Spacer(1, 0.4 * cm),
    ]

    data = [[Paragraph(str(h), head_style) for h in headers]]
    for row in rows:
        data.append([Paragraph('' if v is None else str(v), cell_style) for v in row])

    avail = pagesize[0] - 2 * margin
    col_widths = [avail / ncols] * ncols

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign='CENTER')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    doc.build(elements)
    return response
